from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Request, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, validator
from typing import List, Optional
import uuid
from datetime import datetime, timedelta, timezone
import bcrypt
import jwt
from bson import ObjectId
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import time
import re
import hashlib
from collections import defaultdict, deque

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Environment variables
MONGO_URL = os.getenv('MONGO_URL')
JWT_SECRET = os.getenv('JWT_SECRET', 'your-secret-key-change-this-in-production')

# MongoDB setup
client = AsyncIOMotorClient(MONGO_URL)
db = client.mikel_coffee

# Security Configuration
class SecurityConfig:
    # Rate Limiting
    RATE_LIMIT_REQUESTS = 100  # requests per window
    RATE_LIMIT_WINDOW = 900   # 15 minutes in seconds
    
    # Login Protection
    LOGIN_MAX_ATTEMPTS = 5
    LOGIN_LOCKOUT_TIME = 300  # 5 minutes
    
    # Content Security
    MAX_CONTENT_LENGTH = 10 * 1024 * 1024  # 10MB
    ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.mp4', '.mov', '.avi'}
    
    # Malicious Patterns
    MALICIOUS_PATTERNS = [
        r'<script[^>]*>.*?</script>',  # XSS
        r'javascript:',               # XSS
        r'on\w+\s*=',                # XSS event handlers
        r'union\s+select',           # SQL injection
        r'drop\s+table',             # SQL injection
        r'insert\s+into',            # SQL injection
        r'delete\s+from',            # SQL injection
        r'\.\./\.\.',                # Path traversal
        r'eval\s*\(',                # Code injection
        r'exec\s*\(',                # Code injection
    ]

# Security Classes
class RateLimiter:
    def __init__(self):
        self.requests = defaultdict(deque)
        self.blocked_ips = defaultdict(float)
    
    def is_allowed(self, ip: str) -> bool:
        current_time = time.time()
        
        # Check if IP is currently blocked
        if ip in self.blocked_ips and current_time < self.blocked_ips[ip]:
            return False
        
        # Clean old requests
        window_start = current_time - SecurityConfig.RATE_LIMIT_WINDOW
        while self.requests[ip] and self.requests[ip][0] < window_start:
            self.requests[ip].popleft()
        
        # Check rate limit
        if len(self.requests[ip]) >= SecurityConfig.RATE_LIMIT_REQUESTS:
            # Block IP
            self.blocked_ips[ip] = current_time + SecurityConfig.RATE_LIMIT_WINDOW
            return False
        
        # Add current request
        self.requests[ip].append(current_time)
        return True

class LoginProtection:
    def __init__(self):
        self.failed_attempts = defaultdict(list)
        self.blocked_users = defaultdict(float)
    
    def is_blocked(self, email: str) -> bool:
        current_time = time.time()
        return email in self.blocked_users and current_time < self.blocked_users[email]
    
    def record_failed_attempt(self, email: str):
        current_time = time.time()
        # Clean old attempts (older than lockout time)
        self.failed_attempts[email] = [
            attempt for attempt in self.failed_attempts[email]
            if current_time - attempt < SecurityConfig.LOGIN_LOCKOUT_TIME
        ]
        
        # Add current failed attempt
        self.failed_attempts[email].append(current_time)
        
        # Check if should block
        if len(self.failed_attempts[email]) >= SecurityConfig.LOGIN_MAX_ATTEMPTS:
            self.blocked_users[email] = current_time + SecurityConfig.LOGIN_LOCKOUT_TIME
    
    def record_success(self, email: str):
        # Clear failed attempts on successful login
        if email in self.failed_attempts:
            del self.failed_attempts[email]
        if email in self.blocked_users:
            del self.blocked_users[email]

class InputValidator:
    @staticmethod
    def sanitize_input(text: str) -> str:
        if not text:
            return text
        
        # Remove potentially malicious content
        for pattern in SecurityConfig.MALICIOUS_PATTERNS:
            text = re.sub(pattern, '', text, flags=re.IGNORECASE)
        
        # HTML encode special characters
        text = text.replace('<', '&lt;').replace('>', '&gt;')
        text = text.replace('"', '&quot;').replace("'", '&#x27;')
        
        return text
    
    @staticmethod
    def validate_email(email: str) -> bool:
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))
    
    @staticmethod
    def validate_content_size(content: str) -> bool:
        return len(content.encode('utf-8')) <= SecurityConfig.MAX_CONTENT_LENGTH

# Initialize security components
rate_limiter = RateLimiter()
login_protection = LoginProtection()
input_validator = InputValidator()

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_DELTA = timedelta(days=7)

# Create the main app without a prefix
app = FastAPI(title="Mikel Coffee Employee Registration API")

# Security Middleware
async def security_middleware(request: Request, call_next):
    start_time = time.time()
    
    # Get client IP
    client_ip = request.client.host
    if hasattr(request.state, 'forwarded_for'):
        client_ip = request.state.forwarded_for.split(',')[0].strip()
    
    # Rate limiting check
    if not rate_limiter.is_allowed(client_ip):
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=429,
            content={
                "error": "Rate limit exceeded",
                "message": "Too many requests. Please try again later.",
                "retry_after": SecurityConfig.RATE_LIMIT_WINDOW
            }
        )
    
    # Content length check
    content_length = request.headers.get('content-length')
    if content_length and int(content_length) > SecurityConfig.MAX_CONTENT_LENGTH:
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=413,
            content={"error": "Request entity too large"}
        )
    
    # Process request
    response = await call_next(request)
    
    # Add security headers
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    response.headers["Content-Security-Policy"] = "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; img-src 'self' data: https:; font-src 'self' data:;"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    
    # Add processing time
    process_time = time.time() - start_time
    response.headers["X-Process-Time"] = str(process_time)
    
    # Security logging
    print(f"🔐 SECURITY LOG - Request: {request.method} {request.url.path} from IP: {client_ip} - Time: {process_time:.3f}s")
    
    return response

# Add security middleware
app.middleware("http")(security_middleware)

# Host validation middleware
app.add_middleware(
    TrustedHostMiddleware,
    allowed_hosts=["*"]  # Configure with actual domain in production
)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()

# Position Constants
POSITIONS = [
    "servis personeli",
    "barista", 
    "supervizer",
    "müdür yardımcısı",
    "mağaza müdürü",
    "trainer"
]

# Special roles that admin can assign
SPECIAL_ROLES = [
    "eğitim departmanı"
]

# Role hierarchy for permissions
POSITION_LEVELS = {
    "servis personeli": 1,
    "barista": 2,
    "supervizer": 3,
    "müdür yardımcısı": 4,
    "mağaza müdürü": 5,
    "trainer": 6,
    "eğitim departmanı": 7  # Special role with trainer-level permissions
}

# Define Models
class UserRegister(BaseModel):
    name: str
    surname: str
    email: EmailStr
    password: str
    position: str
    store: Optional[str] = None
    start_date: Optional[str] = None  # İşe giriş tarihi

    class Config:
        json_encoders = {
            ObjectId: str
        }

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class User(BaseModel):
    id: Optional[str] = Field(alias="_id")
    employee_id: str
    name: str
    surname: str
    email: EmailStr
    position: str
    store: str
    start_date: Optional[str] = None  # İşe giriş tarihi
    special_role: Optional[str] = None  # "eğitim departmanı" or None
    is_admin: bool = False
    created_at: datetime = Field(default_factory=datetime.utcnow)
    
    class Config:
        allow_population_by_field_name = True
        json_encoders = {
            ObjectId: str
        }

class UserUpdate(BaseModel):
    special_role: Optional[str] = None

class ExamResult(BaseModel):
    id: Optional[str] = Field(alias="_id")
    employee_id: str
    exam_type: str  # "general" or "management"
    score: float
    max_score: float
    passed: bool
    exam_date: datetime = Field(default_factory=datetime.utcnow)
    created_by: str  # trainer who entered the result
    
    class Config:
        allow_population_by_field_name = True
        json_encoders = {
            ObjectId: str
        }

class ExamResultCreate(BaseModel):
    employee_id: str
    exam_type: str
    score: float
    max_score: float

class Announcement(BaseModel):
    id: Optional[str] = Field(alias="_id")
    title: str
    content: str
    created_by: str  # trainer who created
    created_at: datetime = Field(default_factory=datetime.utcnow)
    is_urgent: bool = False
    likes_count: int = 0
    image_url: Optional[str] = None
    
    class Config:
        allow_population_by_field_name = True
        json_encoders = {
            ObjectId: str
        }

class AnnouncementCreate(BaseModel):
    title: str
    content: str
    is_urgent: bool = False
    image_url: Optional[str] = None

class Token(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: User

# Social Media Models
class Post(BaseModel):
    id: Optional[str] = Field(alias="_id")
    author_id: str
    content: str
    image_url: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    likes_count: int = 0
    comments_count: int = 0
    
    class Config:
        allow_population_by_field_name = True
        json_encoders = {
            ObjectId: str
        }

class PostCreate(BaseModel):
    content: str
    image_url: Optional[str] = None

class Comment(BaseModel):
    id: Optional[str] = Field(alias="_id")
    post_id: str
    author_id: str
    content: str
    created_at: datetime = Field(default_factory=datetime.utcnow)
    
    class Config:
        allow_population_by_field_name = True
        json_encoders = {
            ObjectId: str
        }

class CommentCreate(BaseModel):
    content: str

class Like(BaseModel):
    id: Optional[str] = Field(alias="_id")
    post_id: Optional[str] = None
    announcement_id: Optional[str] = None
    user_id: str
    created_at: datetime = Field(default_factory=datetime.utcnow)
    
    class Config:
        allow_population_by_field_name = True
        json_encoders = {
            ObjectId: str
        }

class Profile(BaseModel):
    id: Optional[str] = Field(alias="_id")
    user_id: str
    profile_image_url: Optional[str] = None
    bio: Optional[str] = None
    updated_at: datetime = Field(default_factory=datetime.utcnow)
    
    class Config:
        allow_population_by_field_name = True
        json_encoders = {
            ObjectId: str
        }

class ProfileUpdate(BaseModel):
    profile_image_url: Optional[str] = None
    bio: Optional[str] = None

# Helper Functions
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + JWT_EXPIRATION_DELTA
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    
    # Convert ObjectId to string for Pydantic
    user["_id"] = str(user["_id"])
    return User(**user)

async def generate_employee_id() -> str:
    # Get the highest employee_id and increment
    last_user = await db.users.find_one(sort=[("employee_id", -1)])
    if last_user:
        last_id = int(last_user["employee_id"])
        new_id = last_id + 1
    else:
        new_id = 1
    return f"{new_id:05d}"

async def create_notifications_for_all_users(title: str, message: str, notification_type: str, related_id: str = None, sender_id: str = None):
    """Tüm kullanıcılara bildirim gönder"""
    # Tüm kullanıcıları al
    users = await db.users.find({}, {"employee_id": 1}).to_list(1000)
    
    # Her kullanıcı için bildirim oluştur
    notifications = []
    for user in users:
        notification_doc = {
            "_id": str(uuid.uuid4()),
            "user_id": user["employee_id"],
            "title": title,
            "message": message,
            "type": notification_type,
            "read": False,
            "created_at": datetime.utcnow(),
            "related_id": related_id,
            "sender_id": sender_id
        }
        notifications.append(notification_doc)
    
    # Toplu olarak ekle
    if notifications:
        await db.notifications.insert_many(notifications)

# Authentication Routes
@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserRegister):
    # Check if position is valid
    if user_data.position not in POSITIONS:
        raise HTTPException(status_code=400, detail="Invalid position")
    
    # Check if email already exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Generate employee ID
    employee_id = await generate_employee_id()
    
    # Hash password
    hashed_password = hash_password(user_data.password)
    
    # Only admins can create admin accounts - no automatic admin
    is_admin = False
    
    # Create user document
    user_doc = {
        "employee_id": employee_id,
        "name": user_data.name,
        "surname": user_data.surname,
        "email": user_data.email,
        "password": hashed_password,
        "position": user_data.position,
        "store": user_data.store if user_data.store else "Belirtilmemiş",
        "start_date": user_data.start_date,  # İşe giriş tarihi
        "special_role": None,  # Default no special role
        "is_admin": is_admin,
        "created_at": datetime.utcnow()
    }
    
    result = await db.users.insert_one(user_doc)
    user_doc["_id"] = str(result.inserted_id)
    
    # Create access token
    access_token = create_access_token({"sub": str(result.inserted_id)})
    
    # Remove password from response
    user_doc.pop("password")
    user = User(**user_doc)
    
    return Token(access_token=access_token, token_type="bearer", user=user)

@api_router.post("/auth/login", response_model=Token)
async def login(user_credentials: UserLogin, request: Request):
    # Input validation and sanitization
    email = input_validator.sanitize_input(user_credentials.email.lower().strip())
    
    if not input_validator.validate_email(email):
        raise HTTPException(status_code=400, detail="Invalid email format")
    
    # Check if user is blocked
    if login_protection.is_blocked(email):
        raise HTTPException(
            status_code=429, 
            detail="Account temporarily locked due to too many failed login attempts. Please try again later."
        )
    
    # Find user
    user = await db.users.find_one({"email": email})
    if not user or not bcrypt.checkpw(user_credentials.password.encode('utf-8'), user['password'].encode('utf-8')):
        # Record failed attempt
        login_protection.record_failed_attempt(email)
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Record successful login
    login_protection.record_success(email)
    
    # Create JWT token with enhanced payload
    payload = {
        'sub': str(user['_id']),
        'email': user['email'],
        'employee_id': user['employee_id'],
        'iat': datetime.utcnow(),
        'exp': datetime.utcnow() + JWT_EXPIRATION_DELTA
    }
    token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
    
    # Log successful login (for security monitoring)
    print(f"🔐 SECURITY LOG - Successful login: {email} from IP: {request.client.host}")
    
    # Convert ObjectId to string for response
    user["_id"] = str(user["_id"])
    
    return {"access_token": token, "token_type": "bearer", "user": User(**user)}

@api_router.get("/auth/me", response_model=User)
async def get_current_user_info(current_user: User = Depends(get_current_user)):
    return current_user

# User Management Routes
@api_router.get("/users", response_model=List[User])
async def get_all_users(current_user: User = Depends(get_current_user)):
    # Admin and education department can view all users
    if (not current_user.is_admin and 
        current_user.special_role != "eğitim departmanı"):
        raise HTTPException(status_code=403, detail="Only admin and education department can view all users")
    
    users = await db.users.find({}, {"password": 0}).to_list(1000)
    # Convert ObjectId to string for each user
    for user in users:
        user["_id"] = str(user["_id"])
    return [User(**user) for user in users]

@api_router.get("/users/me", response_model=User)
async def get_current_user_profile(current_user: User = Depends(get_current_user)):
    """Mevcut kullanıcının profil bilgilerini getir"""
    return current_user

@api_router.put("/users/me")
async def update_current_user_profile(user_update: dict, current_user: User = Depends(get_current_user)):
    """Mevcut kullanıcının profil bilgilerini güncelle"""
    
    # Update user data
    update_data = {}
    if "name" in user_update:
        update_data["name"] = user_update["name"]
    if "surname" in user_update:
        update_data["surname"] = user_update["surname"]
    if "position" in user_update:
        update_data["position"] = user_update["position"]
    if "store" in user_update:
        update_data["store"] = user_update["store"]
    if "password" in user_update and user_update["password"]:
        # Hash new password
        hashed_password = bcrypt.hashpw(user_update["password"].encode('utf-8'), bcrypt.gensalt())
        update_data["password"] = hashed_password
    
    if update_data:
        result = await db.users.update_one(
            {"employee_id": current_user.employee_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="User not found")
    
    # Return updated user
    updated_user = await db.users.find_one({"employee_id": current_user.employee_id})
    if updated_user:
        updated_user["_id"] = str(updated_user["_id"])
        return {"message": "Profile updated successfully", "user": User(**updated_user)}
    
    raise HTTPException(status_code=404, detail="User not found")

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, user_update: dict, current_user: User = Depends(get_current_user)):
    # Admin and education department can update users
    if (not current_user.is_admin and 
        current_user.special_role != "eğitim departmanı"):
        raise HTTPException(status_code=403, detail="Only admin and education department can update users")
    
    # Filter allowed updates
    allowed_updates = {}
    if "name" in user_update:
        allowed_updates["name"] = user_update["name"]
    if "surname" in user_update:
        allowed_updates["surname"] = user_update["surname"]
    if "email" in user_update:
        # Check if email is not already taken
        existing = await db.users.find_one({"email": user_update["email"], "_id": {"$ne": ObjectId(user_id)}})
        if existing:
            raise HTTPException(status_code=400, detail="Email already exists")
        allowed_updates["email"] = user_update["email"]
    if "position" in user_update and user_update["position"] in POSITIONS:
        allowed_updates["position"] = user_update["position"]
    if "store" in user_update:
        allowed_updates["store"] = user_update["store"]
    
    # Only admin can update special roles
    if current_user.is_admin and "special_role" in user_update:
        if user_update["special_role"] in SPECIAL_ROLES or user_update["special_role"] is None:
            allowed_updates["special_role"] = user_update["special_role"]
    
    if not allowed_updates:
        raise HTTPException(status_code=400, detail="No valid updates provided")
    
    result = await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": allowed_updates}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get updated user
    user = await db.users.find_one({"_id": ObjectId(user_id)}, {"password": 0})
    if user:
        user["_id"] = str(user["_id"])
        return User(**user)
    
    raise HTTPException(status_code=404, detail="User not found")

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    # Only admin can delete users
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Only admin can delete users")
    
    # Cannot delete self
    if str(current_user.id) == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"_id": ObjectId(user_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

# Exam Results Routes
@api_router.post("/exam-results", response_model=ExamResult)
async def create_exam_result(exam_data: ExamResultCreate, current_user: User = Depends(get_current_user)):
    # Only trainers and education department can create exam results
    if (current_user.position != "trainer" and 
        current_user.special_role != "eğitim departmanı" and 
        not current_user.is_admin):
        raise HTTPException(status_code=403, detail="Only trainers and education department can enter exam results")
    
    # Verify employee exists
    employee = await db.users.find_one({"employee_id": exam_data.employee_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # For management exam, only barista and supervizer can take it
    if exam_data.exam_type == "management":
        if employee["position"] not in ["barista", "supervizer"]:
            raise HTTPException(status_code=400, detail="Only barista and supervizer can take management exam")
    
    # Calculate if passed (80% is passing score)
    passed = (exam_data.score / exam_data.max_score) >= 0.8
    
    exam_doc = {
        "employee_id": exam_data.employee_id,
        "exam_type": exam_data.exam_type,
        "score": exam_data.score,
        "max_score": exam_data.max_score,
        "passed": passed,
        "exam_date": datetime.utcnow(),
        "created_by": current_user.employee_id
    }
    
    result = await db.exam_results.insert_one(exam_doc)
    exam_doc["_id"] = str(result.inserted_id)
    
    return ExamResult(**exam_doc)

@api_router.get("/exam-results", response_model=List[ExamResult])
async def get_exam_results(employee_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    
    # Regular employees can only see their own results
    if not current_user.is_admin and current_user.position != "trainer":
        query["employee_id"] = current_user.employee_id
    elif employee_id:
        query["employee_id"] = employee_id
    
    results = await db.exam_results.find(query).sort("exam_date", -1).to_list(1000)
    # Convert ObjectId to string for each result
    for result in results:
        result["_id"] = str(result["_id"])
    return [ExamResult(**result) for result in results]

# Announcements Routes
@api_router.post("/announcements", response_model=Announcement)
async def create_announcement(announcement_data: AnnouncementCreate, request: Request, current_user: User = Depends(get_current_user)):
    # Only admin, trainer, or eğitim departmanı can create announcements
    if not (current_user.is_admin or current_user.position == 'trainer' or current_user.special_role == 'eğitim departmanı'):
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Input validation and sanitization
    title = input_validator.sanitize_input(announcement_data.title.strip())
    content = input_validator.sanitize_input(announcement_data.content.strip())
    
    if not title or not content:
        raise HTTPException(status_code=400, detail="Title and content are required")
    
    if not input_validator.validate_content_size(title) or not input_validator.validate_content_size(content):
        raise HTTPException(status_code=413, detail="Content too large")
    
    announcement_id = str(uuid.uuid4())
    announcement_doc = {
        "id": announcement_id,
        "title": title,
        "content": content,
        "is_urgent": announcement_data.is_urgent,
        "created_by": current_user.employee_id,
        "created_at": datetime.utcnow(),
        "likes_count": 0,
        "image_url": input_validator.sanitize_input(announcement_data.image_url) if announcement_data.image_url else None
    }
    
    result = await db.announcements.insert_one(announcement_doc)
    announcement_doc["_id"] = str(result.inserted_id)
    
    # Security logging
    print(f"🔐 SECURITY LOG - Announcement created by: {current_user.email} from IP: {request.client.host}")
    
    # Tüm kullanıcılara bildirim gönder
    await create_notifications_for_all_users(
        title="🔔 Yeni Duyuru",
        message=f"{title[:50]}{'...' if len(title) > 50 else ''}",
        notification_type="announcement",
        related_id=announcement_id,
        sender_id=current_user.employee_id
    )
    
    # Tüm kullanıcılara push notification da gönder
    await send_push_notifications_to_all_users(
        "🔔 Yeni Duyuru - Mikel Coffee",
        f"{title[:100]}{'...' if len(title) > 100 else ''}"
    )
    
    return Announcement(**announcement_doc)

@api_router.get("/announcements", response_model=List[Announcement])
async def get_announcements(current_user: User = Depends(get_current_user)):
    announcements = await db.announcements.find({}).sort("created_at", -1).to_list(1000)
    # Convert ObjectId to string for each announcement
    for announcement in announcements:
        announcement["_id"] = str(announcement["_id"])
    return [Announcement(**announcement) for announcement in announcements]

@api_router.delete("/announcements/{announcement_id}")
async def delete_announcement(announcement_id: str, current_user: User = Depends(get_current_user)):
    # Only admin, trainer, or eğitim departmanı can delete
    if not (current_user.is_admin or current_user.position == 'trainer' or current_user.special_role == 'eğitim departmanı'):
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Try both _id (as ObjectId) and id (as string) fields for compatibility
    query_conditions = [{"id": announcement_id}]  # UUID field
    
    # Try to convert to ObjectId for _id field
    try:
        query_conditions.append({"_id": ObjectId(announcement_id)})
    except:
        # If not a valid ObjectId, also try as string
        query_conditions.append({"_id": announcement_id})
    
    announcement = await db.announcements.find_one({"$or": query_conditions})
    if not announcement:
        raise HTTPException(status_code=404, detail="Announcement not found")
    
    # Delete using the same query conditions
    await db.announcements.delete_one({"$or": query_conditions})
    return {"message": "Announcement deleted"}

# Admin Routes for User Management
@api_router.put("/admin/users/{user_id}/special-role")
async def assign_special_role(user_id: str, user_update: UserUpdate, current_user: User = Depends(get_current_user)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Only admin can assign special roles")
    
    # Validate special role
    if user_update.special_role and user_update.special_role not in SPECIAL_ROLES:
        raise HTTPException(status_code=400, detail="Invalid special role")
    
    # Update user
    result = await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"special_role": user_update.special_role}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get updated user
    user = await db.users.find_one({"_id": ObjectId(user_id)}, {"password": 0})
    if user:
        user["_id"] = str(user["_id"])
        return User(**user)
    
    raise HTTPException(status_code=404, detail="User not found")

@api_router.get("/admin/stores")
async def get_stores(current_user: User = Depends(get_current_user)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Only admin can view store data")
    
    # Get unique stores from users
    stores = await db.users.distinct("store")
    
    # Get statistics for each store
    store_stats = []
    for store in stores:
        employee_count = await db.users.count_documents({"store": store})
        store_stats.append({
            "store": store,
            "employee_count": employee_count
        })
    
    return store_stats

# Excel Export Routes
@api_router.get("/admin/export/users")
async def export_users_excel(current_user: User = Depends(get_current_user)):
    # Only admin and education department can export
    if (not current_user.is_admin and 
        current_user.special_role != "eğitim departmanı"):
        raise HTTPException(status_code=403, detail="Only admin and education department can export data")
    
    # Get all users
    users = await db.users.find({}, {"password": 0}).sort("employee_id", 1).to_list(1000)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mikel Coffee Çalışanlar"
    
    # Headers
    headers = [
        "Sicil No", "Ad", "Soyad", "E-posta", "Pozisyon", 
        "Mağaza", "İşe Giriş Tarihi", "Özel Rol", "Admin", "Kayıt Tarihi"
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add user data
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.get("employee_id", ""))
        ws.cell(row=row, column=2, value=user.get("name", ""))
        ws.cell(row=row, column=3, value=user.get("surname", ""))
        ws.cell(row=row, column=4, value=user.get("email", ""))
        ws.cell(row=row, column=5, value=user.get("position", "").title())
        ws.cell(row=row, column=6, value=user.get("store", "Belirtilmemiş"))
        
        # İşe giriş tarihi
        start_date = user.get("start_date")
        if start_date:
            try:
                # Tarih formatını düzenle
                if isinstance(start_date, str):
                    date_obj = datetime.fromisoformat(start_date)
                    formatted_start_date = date_obj.strftime("%d.%m.%Y")
                else:
                    formatted_start_date = start_date
            except:
                formatted_start_date = start_date
        else:
            formatted_start_date = "Belirtilmemiş"
        ws.cell(row=row, column=7, value=formatted_start_date)
        
        ws.cell(row=row, column=8, value=user.get("special_role", "Yok") or "Yok")
        ws.cell(row=row, column=9, value="Evet" if user.get("is_admin") else "Hayır")
        
        # Format date
        created_at = user.get("created_at")
        if created_at:
            if isinstance(created_at, str):
                try:
                    date_obj = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    formatted_date = date_obj.strftime("%d.%m.%Y %H:%M")
                except:
                    formatted_date = created_at
            else:
                formatted_date = created_at.strftime("%d.%m.%Y %H:%M")
        else:
            formatted_date = "Bilinmiyor"
        
        ws.cell(row=row, column=10, value=formatted_date)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generate filename with date
    filename = f"Mikel_Coffee_Calisanlar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Return as streaming response
    def iter_excel():
        yield excel_buffer.read()
    
    return StreamingResponse(
        iter_excel(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
    
# Test endpoint to migrate old users and make first user admin
@api_router.post("/test/migrate-db")
async def migrate_database():
    # Add store field to users who don't have it
    await db.users.update_many(
        {"store": {"$exists": False}},
        {"$set": {"store": "Bilinmiyor"}}
    )
    
    # Add special_role field to users who don't have it
    await db.users.update_many(
        {"special_role": {"$exists": False}},
        {"$set": {"special_role": None}}
    )
    
    # Make first user admin
    first_user = await db.users.find_one(sort=[("created_at", 1)])
    if first_user:
        await db.users.update_one(
            {"_id": first_user["_id"]},
            {"$set": {"is_admin": True}}
        )
    
# Test endpoint to make specific user admin
@api_router.post("/test/make-admin/{user_email}")
async def make_user_admin(user_email: str):
    user = await db.users.find_one({"email": user_email})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one(
        {"_id": user["_id"]},
        {"$set": {"is_admin": True}}
    )
    
    return {"message": f"User {user_email} is now admin"}

# Statistics Routes (for admin dashboard)
@api_router.get("/stats")
async def get_statistics(current_user: User = Depends(get_current_user)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    total_employees = await db.users.count_documents({})
    
    # Count by position
    position_stats = {}
    for position in POSITIONS:
        count = await db.users.count_documents({"position": position})
        position_stats[position] = count
    
    # Exam statistics
    total_exams = await db.exam_results.count_documents({})
    passed_exams = await db.exam_results.count_documents({"passed": True})
    
    # Management exam eligible count
    management_eligible = await db.users.count_documents({"position": {"$in": ["barista", "supervizer"]}})
    
    return {
        "total_employees": total_employees,
        "position_stats": position_stats,
        "total_exams": total_exams,
        "passed_exams": passed_exams,
        "exam_pass_rate": passed_exams / total_exams if total_exams > 0 else 0,
        "management_exam_eligible": management_eligible
    }

# Social Media Endpoints

@api_router.post("/posts", response_model=Post)
async def create_post(post: PostCreate, request: Request, current_user: User = Depends(get_current_user)):
    # Input validation and sanitization
    content = input_validator.sanitize_input(post.content.strip()) if post.content else ""
    
    if not content and not post.image_url:
        raise HTTPException(status_code=400, detail="Content or media is required")
    
    if content and not input_validator.validate_content_size(content):
        raise HTTPException(status_code=413, detail="Content too large")
    
    post_data = {
        "_id": str(uuid.uuid4()),
        "author_id": current_user.employee_id,
        "content": content,
        "image_url": input_validator.sanitize_input(post.image_url) if post.image_url else None,
        "created_at": datetime.utcnow(),
        "likes_count": 0,
        "comments_count": 0
    }
    
    await db.posts.insert_one(post_data)
    
    # Security logging
    print(f"🔐 SECURITY LOG - Post created by: {current_user.email} from IP: {request.client.host}")
    
    return Post(**post_data)

@api_router.get("/posts", response_model=List[Post])
async def get_posts(current_user: User = Depends(get_current_user)):
    posts = await db.posts.find().sort("created_at", -1).to_list(length=None)
    # Convert ObjectId to string for each post
    for post in posts:
        if "_id" in post:
            post["_id"] = str(post["_id"])
    return [Post(**post) for post in posts]

@api_router.delete("/posts/{post_id}")
async def delete_post(post_id: str, current_user: User = Depends(get_current_user)):
    # Try both _id and id fields for compatibility
    post = await db.posts.find_one({"$or": [{"_id": post_id}, {"id": post_id}]})
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    
    # Only author or admin can delete
    if post["author_id"] != current_user.employee_id and not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Delete using the same identifier
    await db.posts.delete_one({"$or": [{"_id": post_id}, {"id": post_id}]})
    await db.comments.delete_many({"post_id": post_id})
    await db.likes.delete_many({"post_id": post_id})
    return {"message": "Post deleted"}

@api_router.post("/posts/{post_id}/comments", response_model=Comment)
async def create_comment(post_id: str, comment: CommentCreate, request: Request, current_user: User = Depends(get_current_user)):
    # Input validation and sanitization
    content = input_validator.sanitize_input(comment.content.strip())
    
    if not content:
        raise HTTPException(status_code=400, detail="Comment content is required")
    
    if not input_validator.validate_content_size(content):
        raise HTTPException(status_code=413, detail="Comment too large")
    
    # Check if post exists
    post = await db.posts.find_one({"$or": [{"id": post_id}, {"_id": post_id}]})
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    
    comment_data = {
        "id": str(uuid.uuid4()),
        "post_id": post_id,
        "author_id": current_user.employee_id,
        "content": content,
        "created_at": datetime.utcnow()
    }
    
    await db.comments.insert_one(comment_data)
    # Update comment count
    await db.posts.update_one({"$or": [{"id": post_id}, {"_id": post_id}]}, {"$inc": {"comments_count": 1}})
    
    # Security logging
    print(f"🔐 SECURITY LOG - Comment created by: {current_user.email} on post: {post_id} from IP: {request.client.host}")
    
    return Comment(**comment_data)

@api_router.get("/posts/{post_id}/comments", response_model=List[Comment])
async def get_comments(post_id: str, current_user: User = Depends(get_current_user)):
    comments = await db.comments.find({"post_id": post_id}).sort("created_at", 1).to_list(length=None)
    return [Comment(**comment) for comment in comments]

@api_router.post("/posts/{post_id}/like")
async def toggle_post_like(post_id: str, current_user: User = Depends(get_current_user)):
    # Check if post exists - use _id field
    post = await db.posts.find_one({"_id": post_id})
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    
    existing_like = await db.likes.find_one({"post_id": post_id, "user_id": current_user.employee_id})
    
    if existing_like:
        # Unlike
        await db.likes.delete_one({"post_id": post_id, "user_id": current_user.employee_id})
        await db.posts.update_one({"_id": post_id}, {"$inc": {"likes_count": -1}})
        return {"liked": False}
    else:
        # Like
        like_data = {
            "_id": str(uuid.uuid4()),
            "post_id": post_id,
            "user_id": current_user.employee_id,
            "created_at": datetime.utcnow()
        }
        await db.likes.insert_one(like_data)
        await db.posts.update_one({"_id": post_id}, {"$inc": {"likes_count": 1}})
        return {"liked": True}

@api_router.post("/announcements/{announcement_id}/like")
async def toggle_announcement_like(announcement_id: str, current_user: User = Depends(get_current_user)):
    # Check if announcement exists - use ObjectId for MongoDB query
    try:
        announcement = await db.announcements.find_one({"_id": ObjectId(announcement_id)})
    except:
        announcement = None
    
    if not announcement:
        raise HTTPException(status_code=404, detail="Announcement not found")
    
    existing_like = await db.likes.find_one({"announcement_id": announcement_id, "user_id": current_user.employee_id})
    
    if existing_like:
        # Unlike
        await db.likes.delete_one({"announcement_id": announcement_id, "user_id": current_user.employee_id})
        # CRITICAL FIX: Update likes_count in announcements collection
        await db.announcements.update_one({"_id": ObjectId(announcement_id)}, {"$inc": {"likes_count": -1}})
        return {"liked": False}
    else:
        # Like
        like_data = {
            "id": str(uuid.uuid4()),
            "announcement_id": announcement_id,
            "user_id": current_user.employee_id,
            "created_at": datetime.utcnow()
        }
        await db.likes.insert_one(like_data)
        # CRITICAL FIX: Update likes_count in announcements collection
        await db.announcements.update_one({"_id": ObjectId(announcement_id)}, {"$inc": {"likes_count": 1}})
        return {"liked": True}

@api_router.get("/profile", response_model=Profile)
async def get_profile(current_user: User = Depends(get_current_user)):
    profile = await db.profiles.find_one({"user_id": current_user.employee_id})
    if not profile:
        # Create default profile
        profile_data = {
            "_id": str(uuid.uuid4()),
            "user_id": current_user.employee_id,
            "profile_image_url": None,
            "bio": None,
            "updated_at": datetime.utcnow()
        }
        result = await db.profiles.insert_one(profile_data)
        return Profile(**profile_data)
    
    # Convert ObjectId to string
    if "_id" in profile:
        profile["_id"] = str(profile["_id"])
    return Profile(**profile)

@api_router.put("/profile", response_model=Profile)
async def update_profile(profile_update: ProfileUpdate, current_user: User = Depends(get_current_user)):
    update_data = {}
    if profile_update.profile_image_url is not None:
        update_data["profile_image_url"] = profile_update.profile_image_url
    if profile_update.bio is not None:
        update_data["bio"] = profile_update.bio
    update_data["updated_at"] = datetime.utcnow()
    
    # Upsert profile
    await db.profiles.update_one(
        {"user_id": current_user.employee_id},
        {"$set": update_data},
        upsert=True
    )
    
    profile = await db.profiles.find_one({"user_id": current_user.employee_id})
    if profile and "_id" in profile:
        profile["_id"] = str(profile["_id"])
    return Profile(**profile)

@api_router.get("/profiles", response_model=List[Profile])
async def get_all_profiles(current_user: User = Depends(get_current_user)):
    profiles = await db.profiles.find().to_list(length=None)
    # Convert ObjectId to string for each profile
    for profile in profiles:
        if "_id" in profile:
            profile["_id"] = str(profile["_id"])
    return [Profile(**profile) for profile in profiles]

class UserUpdate(BaseModel):
    name: Optional[str] = None
    surname: Optional[str] = None
    position: Optional[str] = None
    store: Optional[str] = None



@api_router.delete("/admin/users/{employee_id}")
async def delete_user(employee_id: str, current_user: User = Depends(get_current_user)):
    # Only admin can delete users
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Only admin can delete users")
    
    # Cannot delete yourself
    if employee_id == current_user.employee_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    # Check if user exists
    user_to_delete = await db.users.find_one({"employee_id": employee_id})
    if not user_to_delete:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Delete user and related data
    await db.users.delete_one({"employee_id": employee_id})
    await db.profiles.delete_many({"user_id": employee_id})
    await db.posts.delete_many({"author_id": employee_id})
    await db.exam_results.delete_many({"employee_id": employee_id})
    await db.likes.delete_many({"user_id": employee_id})
    await db.comments.delete_many({"author_id": employee_id})
    
    return {"message": f"User {employee_id} and all related data deleted successfully"}

class AdminStatusUpdate(BaseModel):
    is_admin: bool
    reason: Optional[str] = None

# Notification Model
class Notification(BaseModel):
    id: Optional[str] = Field(alias="_id")
    user_id: str  # Kime gönderilecek
    title: str
    message: str
    type: str  # "announcement", "system", etc.
    read: bool = False
    created_at: datetime = Field(default_factory=datetime.utcnow)
    related_id: Optional[str] = None  # Duyuru ID'si vs.

    class Config:
        allow_population_by_field_name = True
        populate_by_name = True
        json_encoders = {
            ObjectId: str,
            datetime: lambda v: v.isoformat()
        }

@api_router.put("/admin/users/{employee_id}/admin-status")
async def update_admin_status(
    employee_id: str, 
    admin_update: AdminStatusUpdate, 
    request: Request,
    current_user: User = Depends(get_current_user)
):
    # Only admin can grant or revoke admin privileges
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Only existing admins can grant or revoke admin privileges")
    
    # Cannot modify your own admin status
    if employee_id == current_user.employee_id:
        raise HTTPException(status_code=400, detail="Cannot modify your own admin status")
    
    # Check if target user exists
    target_user = await db.users.find_one({"employee_id": employee_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Input validation
    reason = input_validator.sanitize_input(admin_update.reason) if admin_update.reason else None
    
    # Update admin status
    await db.users.update_one(
        {"employee_id": employee_id},
        {"$set": {"is_admin": admin_update.is_admin}}
    )
    
    # Security logging
    action = "granted" if admin_update.is_admin else "revoked"
    print(f"🔐 SECURITY LOG - Admin privileges {action} for user: {employee_id} by admin: {current_user.email} from IP: {request.client.host}")
    if reason:
        print(f"🔐 SECURITY LOG - Reason: {reason}")
    
    # Get updated user
    updated_user = await db.users.find_one({"employee_id": employee_id})
    updated_user["_id"] = str(updated_user["_id"])
    
    return {
        "message": f"Admin privileges {action} successfully",
        "user": User(**updated_user),
        "action_by": current_user.email,
        "reason": reason
    }

# Notification Endpoints
@api_router.get("/notifications", response_model=List[Notification])
async def get_user_notifications(current_user: User = Depends(get_current_user)):
    """Kullanıcının bildirimlerini getir"""
    notifications = await db.notifications.find(
        {"user_id": current_user.employee_id}
    ).sort("created_at", -1).limit(50).to_list(50)
    
    # Convert ObjectId to string and add id field for Pydantic compatibility
    for notification in notifications:
        notification["id"] = str(notification["_id"])
        notification["_id"] = str(notification["_id"])
    
    return [Notification(**notif) for notif in notifications]

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_as_read(notification_id: str, current_user: User = Depends(get_current_user)):
    """Bildirimi okundu olarak işaretle"""
    result = await db.notifications.update_one(
        {"_id": notification_id, "user_id": current_user.employee_id},
        {"$set": {"read": True}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"message": "Notification marked as read"}

@api_router.get("/notifications/unread-count")
async def get_unread_notifications_count(current_user: User = Depends(get_current_user)):
    """Okunmamış bildirim sayısını getir"""
    count = await db.notifications.count_documents({
        "user_id": current_user.employee_id,
        "read": False
    })
    
    return {"unread_count": count}

# Push Notification Models and Endpoints
class PushSubscription(BaseModel):
    endpoint: str
    keys: dict

@api_router.post("/push/subscribe")
async def subscribe_to_push(subscription: PushSubscription, current_user: User = Depends(get_current_user)):
    """Kullanıcının push notification subscription'ını kaydet"""
    
    # Kullanıcının mevcut subscription'ını güncelle veya yeni oluştur
    await db.push_subscriptions.update_one(
        {"user_id": current_user.employee_id},
        {
            "$set": {
                "user_id": current_user.employee_id,
                "endpoint": subscription.endpoint,
                "keys": subscription.keys,
                "created_at": datetime.utcnow()
            }
        },
        upsert=True
    )
    
    return {"message": "Push subscription saved successfully"}

@api_router.post("/push/send-test")
async def send_test_push(current_user: User = Depends(get_current_user)):
    """Test push notification gönder (admin only)"""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Only admin can send test notifications")
    
    # Bu endpoint test amaçlı - gerçek push notification gönderimi için
    # web push library'si gerekiyor (pywebpush)
    return {"message": "Test push notification would be sent (requires pywebpush implementation)"}

async def send_push_notification_to_user(user_id: str, title: str, body: str):
    """Kullanıcıya push notification gönder"""
    try:
        # Kullanıcının push subscription'ını al
        subscription_doc = await db.push_subscriptions.find_one({"user_id": user_id})
        
        if subscription_doc:
            # Gerçek implementasyon için pywebpush library'si gerekiyor
            # Şimdilik console'a log yazdıralım
            print(f"📱 PUSH NOTIFICATION - User: {user_id}, Title: {title}, Body: {body}")
            print(f"📱 PUSH SUBSCRIPTION: {subscription_doc['endpoint']}")
            
            # TODO: Implement actual push notification sending with pywebpush
            # from pywebpush import webpush, WebPushException
            # webpush(subscription_info=subscription_doc, data=json.dumps({
            #     "title": title,
            #     "body": body
            # }), vapid_private_key="path/to/private_key.pem", vapid_claims={"sub": "mailto:admin@mikelcoffee.com"})
            
        else:
            print(f"📱 NO PUSH SUBSCRIPTION found for user: {user_id}")
            
    except Exception as e:
        print(f"❌ Error sending push notification: {e}")

async def send_push_notifications_to_all_users(title: str, body: str):
    """Tüm kullanıcılara push notification gönder"""
    try:
        # Tüm push subscription'ları al
        subscriptions = await db.push_subscriptions.find({}).to_list(1000)
        
        print(f"📱 SENDING PUSH TO {len(subscriptions)} USERS: {title}")
        
        for subscription in subscriptions:
            await send_push_notification_to_user(subscription["user_id"], title, body)
            
    except Exception as e:
        print(f"❌ Error sending push notifications to all users: {e}")

async def create_notifications_for_all_users(title: str, message: str, notification_type: str, related_id: str = None, sender_id: str = None):
    """Tüm kullanıcılara bildirim oluştur"""
    try:
        # Tüm kullanıcıları al
        users = await db.users.find({}).to_list(1000)
        
        # Her kullanıcı için bildirim oluştur
        notifications = []
        for user in users:
            notification = {
                "user_id": user["employee_id"],
                "title": title,
                "message": message,
                "type": notification_type,
                "read": False,
                "created_at": datetime.utcnow(),
                "related_id": related_id,
                "sender_id": sender_id
            }
            notifications.append(notification)
        
        # Bulk insert for performance
        if notifications:
            await db.notifications.insert_many(notifications)
            print(f"📧 Created {len(notifications)} notifications for all users")
            
    except Exception as e:
        print(f"❌ Error creating notifications for all users: {e}")

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()